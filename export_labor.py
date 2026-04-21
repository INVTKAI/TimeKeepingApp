#!/usr/bin/env python3
"""Generate a formatted Labor Report Excel workbook from JSON records."""
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

def build_report(records_json, output_path):
    records = json.loads(records_json)
    wb = Workbook()

    # --- Detail Sheet ---
    ws = wb.active
    ws.title = 'Labor Detail'
    headers = ['Source','Employee','Emp ID','Craft','Date','Day','Project #','Project Name',
               'Area','Task Code','CWP','FCO/PCR','Type','Hours','Comment']

    hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    hdr_fill = PatternFill('solid', fgColor='1A56DB')
    data_font = Font(name='Arial', size=10)
    hrs_fmt = '#,##0.00;(#,##0.00);"-"'
    ot_font = Font(name='Arial', size=10, color='DC2626', bold=True)
    thin = Side(style='thin', color='D1D5DB')
    border = Border(bottom=thin)
    total_fill = PatternFill('solid', fgColor='F3F4F6')
    total_font = Font(name='Arial', bold=True, size=10)

    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for row_idx, r in enumerate(records, 2):
        vals = [r['source'], r['empName'], r['empId'], r['craft'], r['date'], r['day'],
                r['projectNum'], r['projectName'], r['area'], r['taskCode'],
                r.get('cwp', ''), r['fco'], r['type'], r['hours'], r.get('comment', '')]
        for col_idx, v in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=v)
            c.font = ot_font if r['type'] == 'OT' and col_idx == 13 else data_font
            c.border = border
            if col_idx == 14:
                c.number_format = hrs_fmt
                c.alignment = Alignment(horizontal='right')

    # Totals row with formulas
    tr = len(records) + 2
    ws.cell(row=tr, column=1, value='TOTALS').font = total_font
    for col in range(1, 16):
        ws.cell(row=tr, column=col).fill = total_fill
        ws.cell(row=tr, column=col).font = total_font
        ws.cell(row=tr, column=col).border = Border(top=Side(style='double', color='374151'))

    letter = get_column_letter(14)
    ws.cell(row=tr, column=14).value = f'=SUM({letter}2:{letter}{tr-1})'
    ws.cell(row=tr, column=14).number_format = hrs_fmt
    ws.cell(row=tr, column=14).alignment = Alignment(horizontal='right')

    col_widths = [8,22,8,14,12,10,10,24,10,10,12,10,5,8,24]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f'A1:O{tr}'
    ws.freeze_panes = 'A2'

    # --- Summary Sheet ---
    ws2 = wb.create_sheet('Summary by Project')
    ws2_headers = ['Project #','Project Name','ST Hours','OT Hours','Total Hours']
    for col_idx, h in enumerate(ws2_headers, 1):
        c = ws2.cell(row=1, column=col_idx, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    proj_data = {}
    for r in records:
        key = r['projectNum']
        if key not in proj_data:
            proj_data[key] = {'name': r['projectName'], 'st': 0, 'ot': 0}
        pd = proj_data[key]
        if r['type'] == 'ST':
            pd['st'] += r['hours']
        else:
            pd['ot'] += r['hours']

    for row_idx, (pnum, pd) in enumerate(sorted(proj_data.items()), 2):
        total_hrs = pd['st'] + pd['ot']
        vals = [pnum, pd['name'], pd['st'], pd['ot'], total_hrs]
        for col_idx, v in enumerate(vals, 1):
            c = ws2.cell(row=row_idx, column=col_idx, value=v)
            c.font = data_font
            c.border = border
            if col_idx in (3, 4, 5):
                c.number_format = hrs_fmt

    ws2_widths = [10, 26, 10, 10, 10]
    for i, w in enumerate(ws2_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    print(json.dumps({"status": "success", "path": output_path, "records": len(records)}))

if __name__ == '__main__':
    build_report(sys.argv[1], sys.argv[2])
